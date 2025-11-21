"""
Export service for KPI and reports.
"""
from typing import Dict, List
from io import BytesIO
from fastapi.responses import StreamingResponse, Response
import csv

from services.kpi import KPIService
from google_sheets import sheets_client


async def export_kpi_data(format: str = "csv") -> StreamingResponse:
    """Export KPI data in the specified format."""
    kpi_service = KPIService()
    seller_stats = await kpi_service.get_seller_stats()
    
    if format == "csv":
        return await _export_csv(seller_stats)
    elif format == "excel":
        return await _export_excel(seller_stats)
    elif format == "pdf":
        return await _export_pdf(seller_stats)
    else:
        raise ValueError(f"Unsupported format: {format}")


async def _export_csv(seller_stats: Dict) -> StreamingResponse:
    """Export KPI data as CSV."""
    output = BytesIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "Seller Name",
        "Total Leads",
        "Call #1 Completion %",
        "Call #2 Completion %",
        "Call #3 Completion %",
        "Follow-ups Completed",
        "First Lessons Scheduled",
        "Attendance Rate %",
        "Lead → First Lesson %",
        "First Lesson → Sale %",
        "Lead → Sale %"
    ])
    
    # Data rows
    for seller_name, stats in seller_stats.items():
        writer.writerow([
            seller_name,
            stats.get("total_leads", 0),
            stats.get("call1_completion_rate", 0),
            stats.get("call2_completion_rate", 0),
            stats.get("call3_completion_rate", 0),
            stats.get("followups_completed", 0),
            stats.get("first_lessons_scheduled", 0),
            stats.get("attendance_rate", 0),
            stats.get("lead_to_first_lesson_rate", 0),
            stats.get("first_lesson_to_sale_rate", 0),
            stats.get("lead_to_sale_rate", 0),
        ])
    
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=kpi_report.csv"}
    )


async def _export_excel(seller_stats: Dict) -> StreamingResponse:
    """Export KPI data as Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "KPI Report"
        
        # Header
        headers = [
            "Seller Name",
            "Total Leads",
            "Call #1 Completion %",
            "Call #2 Completion %",
            "Call #3 Completion %",
            "Follow-ups Completed",
            "First Lessons Scheduled",
            "Attendance Rate %",
            "Lead → First Lesson %",
            "First Lesson → Sale %",
            "Lead → Sale %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, (seller_name, stats) in enumerate(seller_stats.items(), 2):
            ws.cell(row=row, column=1, value=seller_name)
            ws.cell(row=row, column=2, value=stats.get("total_leads", 0))
            ws.cell(row=row, column=3, value=stats.get("call1_completion_rate", 0))
            ws.cell(row=row, column=4, value=stats.get("call2_completion_rate", 0))
            ws.cell(row=row, column=5, value=stats.get("call3_completion_rate", 0))
            ws.cell(row=row, column=6, value=stats.get("followups_completed", 0))
            ws.cell(row=row, column=7, value=stats.get("first_lessons_scheduled", 0))
            ws.cell(row=row, column=8, value=stats.get("attendance_rate", 0))
            ws.cell(row=row, column=9, value=stats.get("lead_to_first_lesson_rate", 0))
            ws.cell(row=row, column=10, value=stats.get("first_lesson_to_sale_rate", 0))
            ws.cell(row=row, column=11, value=stats.get("lead_to_sale_rate", 0))
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=kpi_report.xlsx"}
        )
    except ImportError:
        raise ValueError("openpyxl is required for Excel export. Install with: pip install openpyxl")


async def _export_pdf(seller_stats: Dict) -> StreamingResponse:
    """Export KPI data as PDF."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph("KPI Report", styles["Title"]))
        elements.append(Spacer(1, 12))
        
        # Table data
        data = [["Seller Name", "Total Leads", "Call #1 %", "Call #2 %", "Call #3 %", "Follow-ups", "First Lessons", "Attendance %", "Lead→Lesson %", "Lesson→Sale %", "Lead→Sale %"]]
        
        for seller_name, stats in seller_stats.items():
            data.append([
                seller_name,
                str(stats.get("total_leads", 0)),
                f"{stats.get('call1_completion_rate', 0):.1f}%",
                f"{stats.get('call2_completion_rate', 0):.1f}%",
                f"{stats.get('call3_completion_rate', 0):.1f}%",
                str(stats.get("followups_completed", 0)),
                str(stats.get("first_lessons_scheduled", 0)),
                f"{stats.get('attendance_rate', 0):.1f}%",
                f"{stats.get('lead_to_first_lesson_rate', 0):.1f}%",
                f"{stats.get('first_lesson_to_sale_rate', 0):.1f}%",
                f"{stats.get('lead_to_sale_rate', 0):.1f}%",
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=kpi_report.pdf"}
        )
    except ImportError:
        raise ValueError("reportlab is required for PDF export. Install with: pip install reportlab")

